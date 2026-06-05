"""
XLSX-Export für Businessplan-Finanzmodell.
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.services.businessplan.models import BusinessPlanResult


def export_xlsx(result: BusinessPlanResult) -> bytes:
    """Erzeugt ein Excel-Workbook mit Finanzmodell, Prüflogik, Förderprogrammen."""
    wb = Workbook()

    # --- Sheet 1: Finanzmodell ---
    ws1 = wb.active
    ws1.title = "Finanzmodell"
    _write_financials_sheet(ws1, result)

    # --- Sheet 2: Prüflogik ---
    ws2 = wb.create_sheet("Prüflogik")
    _write_checks_sheet(ws2, result)

    # --- Sheet 3: Fördermittel ---
    ws3 = wb.create_sheet("Fördermittel")
    _write_funding_sheet(ws3, result)

    # --- Sheet 4: Score ---
    ws4 = wb.create_sheet("Score")
    _write_score_sheet(ws4, result)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def _header_style(cell, color: str = "D9EAF7") -> None:
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor=color)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _autosize(ws, max_width: int = 35) -> None:
    for col in ws.columns:
        col_letter = col[0].column_letter
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def _write_financials_sheet(ws, result: BusinessPlanResult) -> None:
    headers = [
        "Jahr",
        "Kunden",
        "Wiederkehrender Umsatz (€)",
        "Setup-Umsatz (€)",
        "Gesamtumsatz (€)",
        "Fixkosten (€)",
        "Marketing (€)",
        "Entwicklung (€)",
        "Gesamtkosten (€)",
        "Ergebnis v. St. (€)",
    ]
    ws.append(headers)
    for cell in ws[1]:
        _header_style(cell, "D9EAF7")

    for f in result.financials:
        ws.append([
            f.year, f.customers,
            f.recurring_revenue, f.setup_revenue, f.total_revenue,
            f.fixed_costs, f.marketing, f.development, f.total_costs,
            f.profit_before_tax,
        ])
    _autosize(ws)


def _write_checks_sheet(ws, result: BusinessPlanResult) -> None:
    headers = ["Bereich", "Status", "Befund", "Empfehlung"]
    ws.append(headers)
    for cell in ws[1]:
        _header_style(cell, "E2F0D9")

    for c in result.checks:
        ws.append([c.area, c.status, c.finding, c.recommendation])
    _autosize(ws, max_width=60)


def _write_funding_sheet(ws, result: BusinessPlanResult) -> None:
    headers = ["Programm", "Passung", "Warum", "Nächster Schritt", "Region", "URL"]
    ws.append(headers)
    for cell in ws[1]:
        _header_style(cell, "FCE4D6")

    for fm in result.funding:
        ws.append([fm.name, fm.fit, fm.why, fm.next_step, fm.region or "", fm.url or ""])
    _autosize(ws, max_width=60)


def _write_score_sheet(ws, result: BusinessPlanResult) -> None:
    ws.append(["Kriterium", "Score"])
    for cell in ws[1]:
        _header_style(cell, "FFF2CC")

    for label, value in [
        ("Businessplan-Reifegrad", result.scores.business_plan_maturity),
        ("Bankenfähigkeit", result.scores.bankability),
        ("Förderfähigkeit", result.scores.fundability),
        ("Investorenfähigkeit", result.scores.investability),
    ]:
        ws.append([label, f"{value}/100"])
    _autosize(ws)
